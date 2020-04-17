from openpyxl import Workbook, load_workbook
import os
import numpy as np
import matplotlib.pyplot as plt
from matplotlib import rcParams
import matplotlib.figure as figure
import scipy.stats as scip

fig, (ax1, ax2, ax3) = plt.subplots(3,1, figsize=(10,25))

rcParams['font.family'] = 'serif'
rcParams['font.serif'] = ['Times New Roman']
rcParams.update({'font.size': 15})

############################## FIGURE 4A ##############################
x = [5, 25, 50, 75, 100, 125, 150, 175, 200, 225, 250]

##### Initialize means across time bins
SVM_acc_mean = np.empty(len(x))
SNN_acc_mean = np.empty(len(x))
DNN_acc_mean = np.empty(len(x))

##### Initialize stds aross time bins
SVM_acc_sem = np.empty(len(x))
SNN_acc_sem = np.empty(len(x))
DNN_acc_sem = np.empty(len(x))

for bin in range(len(x)):
    ##### INPUT Required Below: #####
    dir_to_unt = 'C:/Users/bioel/PycharmProjects/untitled4/Time_binning_figure/'
    folder_to_plot = 'All V1_' + str(x[bin]) + '_ms'
    #################################
    folder_dir = os.listdir(dir_to_unt + folder_to_plot)

    wb = Workbook()
    SVM_acc = np.empty(len(folder_dir))
    SNN_acc = np.empty(len(folder_dir))
    DNN_acc = np.empty(len(folder_dir))

    for single_session in range(len(folder_dir)):
        one = folder_dir[single_session]

        # change this eventually when excel files are just sessions:
        file = dir_to_unt +folder_to_plot + '/'+ str(one)
        wbx = load_workbook(filename=file)

        SVM_sheet = wbx.get_sheet_by_name('SVM')
        SNN_sheet = wbx.get_sheet_by_name('SNN')
        DNN_sheet = wbx.get_sheet_by_name('DNN')

        cell_SVM = SVM_sheet['B2']
        SVM_acc[single_session] = cell_SVM.value

        cell_SNN = SNN_sheet['G2']
        SNN_acc[single_session] = cell_SNN.value

        cell_DNN = DNN_sheet['G2']
        DNN_acc[single_session] = cell_DNN.value

        continue

    SVM_acc_mean[bin] = np.mean(SVM_acc)
    SNN_acc_mean[bin] = np.mean(SNN_acc)
    DNN_acc_mean[bin] = np.mean(DNN_acc)

    ####################################

    SVM_acc_sem[bin] = scip.sem(SVM_acc)
    SNN_acc_sem[bin] = scip.sem(SNN_acc)
    DNN_acc_sem[bin] = scip.sem(DNN_acc)

    continue

x_vals = np.array(x)

y1 = SVM_acc_mean*100
y2 = SNN_acc_mean*100
y3 = DNN_acc_mean*100

yerr1 = SVM_acc_sem*100
yerr2 = SNN_acc_sem*100
yerr3 = DNN_acc_sem*100

data = [y1, y2, y3]
data_sem = [yerr1,yerr2,yerr3]
columns = x_vals
rows = ['%d year' % x for x in (100, 75, 50)]

ax1.axis([0, 255, 0, 100])
#plt.title('Machine Learning-Based Decoding Accuracies')

colors = plt.cm.BuPu(np.linspace(0.2, 0.7, len(rows)))
ecols = plt.cm.BuPu(np.linspace(0.5,1.0, len(rows)))
n_rows = len(data)
cell_text = []
colors = colors[::-1]

for row in range(n_rows):
    #plt.bar(index, data[row], bar_width, yerr=data_sem[row], bottom=y_offset, color=colors[row])
    #y_offset = y_offset + data[row]
    cell_text.append(['%1.1f' % (x) for x in data[row]])

for row in range(len(data_sem)):
    for col in range(len(data_sem[0])):
        sem_inject = str('%1.1f' % data_sem[row][col])
        cell_text[row][col] = cell_text[row][col] + ' ± ' + sem_inject
        continue
    continue
#ax1.figure(figsize=(12,9))

ax1.errorbar(x_vals,y1,yerr=[yerr1/2, yerr1/2], color=colors[2], ecolor=colors[2], capsize=5.0)
ax1.errorbar(x_vals,y2,yerr=[yerr2/2, yerr2/2], color=colors[1], ecolor=colors[1], capsize=5.0)
ax1.errorbar(x_vals,y3,yerr=[yerr3/2, yerr3/2], color=colors[0],ecolor=colors[0], capsize=5.0)
cell_text.reverse()

# convert time bin durations to str to append ms
columns_ms = []
for i in range(len(columns)):
    column_txt = str(columns[i]) + ' ms'
    columns_ms = np.append(columns_ms, [column_txt])

the_table = ax1.table(cellText=cell_text,
                      rowLabels=['DNN','SNN','SVM'],
                      rowColours=colors,
                      colLabels=columns_ms,
                      cellLoc='center',
                      colWidths=[.0909,.0909,.0909,.0909,.0909,.0909,.0909,.0909,.0909,.0909,.0909],
                      rowLoc='center',
                      colLoc='center',
                      loc='bottom',
                      )

the_table.scale(1,4.0)
#the_table.auto_set_font_size(False)
#the_table.set_fontsize(15)
#the_table.title('Machine Learning-Based Visual Decoding Accuracies per Time Bin Duration')
#plt.axvline([20,40,60])


font = {'fontname': 'Times New Roman',
        'size': 15}
ax1.set_ylabel('Decoding Accuracy (%)', fontdict=font)
ax1.set_ylim(0, 100,)
ax1.set_xticks([])
ax1.text(-.1,1.1,s='(a)', transform=ax1.transAxes, size=15)
#plt.tight_layout()
#plt.legend()
#plt.title('Visual Decoding Accuracies Versus Time Bin Duration For All V1 Units', fontdict=font)

############################## END FIGURE 4A ##############################

############################## FIGURE 4B ##############################
areas = ['VISp','VISrl','VISam', 'VISl', 'VISpm', 'VISal', 'LGN', 'LP', 'CA1', 'CA3']

##### Initialize means across time bins
SVM_acc_mean = np.empty(len(areas))
SNN_acc_mean = np.empty(len(areas))
DNN_acc_mean = np.empty(len(areas))

##### Initialize stds aross time bins
SVM_acc_sem = np.empty(len(areas))
SNN_acc_sem = np.empty(len(areas))
DNN_acc_sem = np.empty(len(areas))

for bin in range(len(areas)):
    ##### INPUT Required Below: #####
    file_header = 'C:/Users/bioel/PycharmProjects/untitled4/Local_anat_figure/'
    folder_to_plot = areas[bin] + '_' + str(50) + '_ms'
    #################################
    folder_dir = os.listdir('C:/Users/bioel/PycharmProjects/untitled4/Local_anat_figure/' + folder_to_plot)
    wb = Workbook()
    SVM_acc = np.empty(len(folder_dir))
    SNN_acc = np.empty(len(folder_dir))
    DNN_acc = np.empty(len(folder_dir))

    for single_session in range(len(folder_dir)):
        one = folder_dir[single_session]
        # change this eventually when excel files are just sessions:
        file = file_header + folder_to_plot + '/' + str(one)
        wbx = load_workbook(filename=file)

        SVM_sheet = wbx.get_sheet_by_name('SVM')
        SNN_sheet = wbx.get_sheet_by_name('SNN')
        DNN_sheet = wbx.get_sheet_by_name('DNN')

        cell_SVM = SVM_sheet['B2']
        SVM_acc[single_session] = cell_SVM.value

        cell_SNN = SNN_sheet['G2']
        SNN_acc[single_session] = cell_SNN.value

        cell_DNN = DNN_sheet['G2']
        DNN_acc[single_session] = cell_DNN.value

        continue

    SVM_acc_mean[bin] = np.mean(SVM_acc)
    SNN_acc_mean[bin] = np.mean(SNN_acc)
    DNN_acc_mean[bin] = np.mean(DNN_acc)

    ####################################

    SVM_acc_sem[bin] = scip.sem(SVM_acc)
    SNN_acc_sem[bin] = scip.sem(SNN_acc)
    DNN_acc_sem[bin] = scip.sem(DNN_acc)

    continue

a = SVM_acc_mean*100
b = SNN_acc_mean*100
c = DNN_acc_mean*100

yerr1 = SVM_acc_sem*100
yerr2 = SNN_acc_sem*100
yerr3 = DNN_acc_sem*100

### Up is me ###

data = [a,b,c]
print(data)
data_sem = [yerr1,yerr2,yerr3]
print(data_sem)
columns = areas
rows = ['%d year' % x for x in (100, 75, 50)]
#values = np.arange(0, 100, 10)
#value_increment = 5

# Get some pastel shades for the colors
colors = plt.cm.BuPu(np.linspace(0.2, 0.7, len(rows)))
ecols = plt.cm.BuPu(np.linspace(0.5,1.0, len(rows)))
n_rows = len(data)



# Initialize the vertical-offset for the stacked bar chart.
y_offset = np.zeros(len(columns))
# Plot bars and create text labels for the table
cell_text = []
colors = colors[::-1]


#x = np.arange(len(areas))
bar_width = 0.2
width_between = 0.2
r1 = np.arange(len(a))
r2 = [x + width_between for x in r1]
r3 = [x + width_between for x in r2]

#ax2.figure(figsize=(12,9))

ax2.bar(r1, data[0], bar_width,align='edge', yerr=[data_sem[0]/2,data_sem[0]/2], ecolor=ecols[2], capsize=3.0,bottom=y_offset, color=colors[2])
ax2.bar(r2, data[1], bar_width,align='edge', yerr=[data_sem[1]/2,data_sem[1]/2],ecolor=ecols[2],capsize=3.0, bottom=y_offset, color=colors[1])
ax2.bar(r3, data[2], bar_width, align='edge',yerr=[data_sem[2]/2,data_sem[2]/2], ecolor=ecols[2],capsize=3.0, bottom=y_offset, color=colors[0])

#ax2.bar(r1, data[2]-data[1], bar_width, yerr=data_sem[2], ecolor=ecols[0], bottom=data[1], color=colors[0])
#ax2.bar(r2, data[1]-data[0], bar_width, yerr=data_sem[1],ecolor=ecols[1], bottom=data[0], color=colors[1])
#ax2.bar(r3, data[0], bar_width, yerr=data_sem[0], ecolor=ecols[2],bottom=y_offset, color=colors[2])


for row in range(n_rows):
    #ax2.bar(index, data[row], bar_width, yerr=data_sem[row], bottom=y_offset, color=colors[row])
    y_offset = y_offset + data[row]
    cell_text.append(['%1.1f' % (x) for x in data[row]])
    print(cell_text)
# Reverse colors and text labels to display the last value at the top.

# He He He!
for row in range(len(data_sem)):
    for col in range(len(data_sem[0])):
        sem_inject = str('%1.1f' % data_sem[row][col])
        cell_text[row][col] = cell_text[row][col] + ' ± ' + sem_inject
        continue
    continue


cell_text.reverse()
print(cell_text)


# Add a table at the bottom of the axes
the_table = ax2.table(cellText=cell_text,
                      rowLabels=['DNN','SNN','SVM'],
                      rowColours=colors,
                      colLabels=columns,
                        cellLoc='center',
                      colWidths=[.1,.1,.1,.1,.1,.1,.1,.1,.1,.1],
                      rowLoc='center',
                      colLoc='center',
                      loc='bottom')

#ax2.yticks(values, ['%d' % val for val in values])
# Adjust layout to make room for the table:
#ax2.subplots_adjust(left=0.2, bottom=0.2)
the_table.scale(1,4.0)
#the_table.auto_set_font_size(False)
#the_table.set_fontsize(15)
font = {'fontname': 'Times New Roman',
        'size': 15}
ax2.set_ylabel('Decoding Accuracy (%)', fontdict=font)
ax2.set_xlim(-.2,9.8)
ax2.set_ylim(0,100)
ax2.set_xticks([])
#ax2.title('Visual Decoding Accuracies of Individual Subregions (fixed 50 ms time bins)')
ax2.text(-.1,1.1,s='(b)', transform=ax2.transAxes, size=15)
### SIGS ###

#first sig
y_max = 50
y_min = 5

ax2.annotate("", xy=(0.1, y_max), xycoords='data',
           xytext=(5.1, y_max), textcoords='data',
           arrowprops=dict(arrowstyle="-", ec='#aaaaaa',
                           connectionstyle="bar,fraction=0.04"))

ax2.text(2.6, y_max + 13.5, 'n.s.',
       horizontalalignment='center',
       verticalalignment='center')

#second sig
y_max = 60
y_min =15

ax2.annotate("", xy=(0.3, y_max), xycoords='data',
           xytext=(5.3, y_max), textcoords='data',
           arrowprops=dict(arrowstyle="-", ec='#aaaaaa',
                           connectionstyle="bar,fraction=0.04"))

ax2.text(2.8, y_max + abs(y_max - y_min)*0.21, '*',
       horizontalalignment='center',
       verticalalignment='center')

#third sig
y_max = 70
y_min = 25

ax2.annotate("", xy=(0.5, y_max), xycoords='data',
           xytext=(5.5, y_max), textcoords='data',
           arrowprops=dict(arrowstyle="-", ec='#aaaaaa',
                           connectionstyle="bar,fraction=0.04"))

ax2.text(3.0, y_max + abs(y_max - y_min)*0.21, '*',
       horizontalalignment='center',
       verticalalignment='center')

#fourth sig
y_max = 45
y_min = 0

ax2.annotate("", xy=(6.1, y_max), xycoords='data',
           xytext=(7.1, y_max), textcoords='data',
           arrowprops=dict(arrowstyle="-", ec='#aaaaaa',
                           connectionstyle="bar,fraction=0.15"))

ax2.text(6.6, y_max + abs(y_max - y_min)*0.16, '*',
       horizontalalignment='center',
       verticalalignment='center')

#fifth sig
y_max = 55
y_min = 10

ax2.annotate("", xy=(6.3, y_max), xycoords='data',
           xytext=(7.3, y_max), textcoords='data',
           arrowprops=dict(arrowstyle="-", ec='#aaaaaa',
                           connectionstyle="bar,fraction=0.15"))

ax2.text(6.8, y_max + abs(y_max - y_min)*0.16, '**',
       horizontalalignment='center',
       verticalalignment='center')

#sixth sig
y_max = 65
y_min = 20

ax2.annotate("", xy=(6.5, y_max), xycoords='data',
           xytext=(7.5, y_max), textcoords='data',
           arrowprops=dict(arrowstyle="-", ec='#aaaaaa',
                           connectionstyle="bar,fraction=0.15"))

ax2.text(7, y_max + abs(y_max - y_min)*0.16, '**',
       horizontalalignment='center',
       verticalalignment='center')

#############################END 4B ######################################

############################## FIGURE 4C ##############################
folders_to_grab = ['All Subcortical', 'All V1 + LGN', 'All Cortical', 'All Units']

##### Initialize means across time bins
SVM_acc_mean = np.empty(len(folders_to_grab))
SNN_acc_mean = np.empty(len(folders_to_grab))
DNN_acc_mean = np.empty(len(folders_to_grab))

##### Initialize stds aross time bins
SVM_acc_sem = np.empty(len(folders_to_grab))
SNN_acc_sem = np.empty(len(folders_to_grab))
DNN_acc_sem = np.empty(len(folders_to_grab))

for bin in range(len(folders_to_grab)):
    ##### INPUT Required Below: #####
    folder_to_plot = folders_to_grab[bin] + '_' + str(50) + '_ms'
    #################################
    folder_dir = os.listdir('C:/Users/bioel/PycharmProjects/untitled4/' + folder_to_plot)

    wb = Workbook()
    SVM_acc = np.empty(len(folder_dir))
    SNN_acc = np.empty(len(folder_dir))
    DNN_acc = np.empty(len(folder_dir))

    for single_session in range(len(folder_dir)):
        one = folder_dir[single_session]

        # change this eventually when excel files are just sessions:
        file = folder_to_plot + '/' + str(one)
        wbx = load_workbook(filename=file)

        SVM_sheet = wbx.get_sheet_by_name('SVM')
        SNN_sheet = wbx.get_sheet_by_name('SNN')
        DNN_sheet = wbx.get_sheet_by_name('DNN')

        cell_SVM = SVM_sheet['B2']
        SVM_acc[single_session] = cell_SVM.value

        cell_SNN = SNN_sheet['G2']
        SNN_acc[single_session] = cell_SNN.value

        cell_DNN = DNN_sheet['G2']
        DNN_acc[single_session] = cell_DNN.value

        continue


    SVM_acc_mean[bin] = np.mean(SVM_acc)
    SNN_acc_mean[bin] = np.mean(SNN_acc)
    DNN_acc_mean[bin] = np.mean(DNN_acc)

    ####################################

    SVM_acc_sem[bin] = scip.sem(SVM_acc)
    SNN_acc_sem[bin] = scip.sem(SNN_acc)
    DNN_acc_sem[bin] = scip.sem(DNN_acc)

    continue

a = SVM_acc_mean*100
b = SNN_acc_mean*100
c = DNN_acc_mean*100

yerr1 = SVM_acc_sem*100
yerr2 = SNN_acc_sem*100
yerr3 = DNN_acc_sem*100

### Up is me ###

data = [a,b,c]
data_sem = [yerr1,yerr2,yerr3]
print(data)
print(data_sem)
columns = folders_to_grab
rows = ['%d year' % x for x in (100, 75, 50)]
#values = np.arange(0, 100, 10)
#value_increment = 5

# Get some pastel shades for the colors
colors = plt.cm.BuPu(np.linspace(0.2, 0.7, len(rows)))
ecols = plt.cm.BuPu(np.linspace(0.5,1.0, len(rows)))
n_rows = len(data)



# Initialize the vertical-offset for the stacked bar chart.
y_offset = np.zeros(len(columns))
# Plot bars and create text labels for the table
cell_text = []
colors = colors[::-1]


#x = np.arange(len(areas))
bar_width = 0.2
width_between = 0.2
r1 = np.arange(len(a))
r2 = [x + width_between for x in r1]
r3 = [x + width_between for x in r2]

#ax3.figure(figsize=(12,9))
ax3.bar(r1, data[0], bar_width,align='edge', yerr=[data_sem[0]/2,data_sem[0]/2], ecolor=ecols[2], capsize=5.0,bottom=y_offset, color=colors[2])
ax3.bar(r2, data[1], bar_width,align='edge', yerr=[data_sem[1]/2,data_sem[1]/2],ecolor=ecols[2],capsize=5.0, bottom=y_offset, color=colors[1])
ax3.bar(r3, data[2], bar_width, align='edge',yerr=[data_sem[2]/2,data_sem[2]/2], ecolor=ecols[2],capsize=5.0, bottom=y_offset, color=colors[0])

#ax3.bar(r1, data[2]-data[1], bar_width, yerr=data_sem[2], ecolor=ecols[0], bottom=data[1], color=colors[0])
#ax3.bar(r2, data[1]-data[0], bar_width, yerr=data_sem[1],ecolor=ecols[1], bottom=data[0], color=colors[1])
#ax3.bar(r3, data[0], bar_width, yerr=data_sem[0], ecolor=ecols[2],bottom=y_offset, color=colors[2])


for row in range(n_rows):
    #ax3.bar(index, data[row], bar_width, yerr=data_sem[row], bottom=y_offset, color=colors[row])
    y_offset = y_offset + data[row]
    cell_text.append(['%1.2f' % (x) for x in data[row]])

# Reverse colors and text labels to display the last value at the top.
# He He He!
for row in range(len(data_sem)):
    for col in range(len(data_sem[0])):
        sem_inject = str('%1.2f' % data_sem[row][col])
        cell_text[row][col] = cell_text[row][col] + ' ± ' + sem_inject
        continue
    continue


cell_text.reverse()
print(cell_text)

# Add a table at the bottom of the axes
the_table = ax3.table(cellText=cell_text,
                      rowLabels=['DNN','SNN','SVM'],
                      rowColours=colors,
                      colLabels=columns,
                        cellLoc='center',
                      colWidths=[.25,.25,.25,.25],
                      rowLoc='center',
                      colLoc='center',
                      loc='bottom')

#ax3.yticks(values, ['%d' % val for val in values])
# Adjust layout to make room for the table:

font = {'fontname': 'Times New Roman',
        'size': 15}

#ax3.subplots_adjust(left=0.2, bottom=0.2)
the_table.scale(1,4)
#the_table.auto_set_font_size(False)
#the_table.set_fontsize(10)
ax3.set_ylabel('Decoding Accuracy (%)', fontdict=font)
ax3.set_ylim(0, 100)
ax3.set_xlim(-.2, 3.8)
ax3.set_xticks([])
#ax3.title('Visual Decoding Accuracies of Grouped Regions (fixed 50 ms time bins)')
ax3.text(-.1,1.1,s='(c)', transform=ax3.transAxes, size=15)

#first sig
y_max = 70
y_min = 0

ax3.annotate("", xy=(0.1, y_max), xycoords='data',
           xytext=(1.1, y_max), textcoords='data',
           arrowprops=dict(arrowstyle="-", ec='#aaaaaa',
                           connectionstyle="bar,fraction=0.075"))

ax3.text(0.6, y_max + abs(y_max - y_min)*0.125, '***',
       horizontalalignment='center',
       verticalalignment='center')


#end first sig
#second sig
y_min=10
y_max=77.5
ax3.annotate("", xy=(.3, y_max), xycoords='data',
           xytext=(1.3, y_max), textcoords='data',
           arrowprops=dict(arrowstyle="-", ec='#aaaaaa',
                           connectionstyle="bar,fraction=0.075"))

ax3.text(0.8, y_max + abs(y_max - y_min)*0.125, '***',
       horizontalalignment='center',
       verticalalignment='center')

#end second sig
#third sig
y_min=20
y_max=85
ax3.annotate("", xy=(.5, y_max), xycoords='data',
           xytext=(1.5, y_max), textcoords='data',
           arrowprops=dict(arrowstyle="-", ec='#aaaaaa',
                           connectionstyle="bar,fraction=0.075"))

ax3.text(1.0, y_max + abs(y_max - y_min)*0.125, '***',
       horizontalalignment='center',
       verticalalignment='center')
#end third sig

fig.tight_layout(pad=1.0)
plt.show()