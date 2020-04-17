from openpyxl import Workbook, load_workbook
import os
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.figure as figure
import matplotlib.axes as ax
import scipy.stats as scip


############################## FIGURE 4C ##############################
folders_to_grab = ['All Subcortical', 'All V1 + LGN', 'All Cortical', 'All']


##### Initialize Stats #####
SVM_acc_stats = np.empty([31,len(folders_to_grab)])
SNN_acc_stats = np.empty([31,len(folders_to_grab)])
DNN_acc_stats = np.empty([31,len(folders_to_grab)])



##### All SubCortical #####
folder_to_plot = folders_to_grab[0] + '_' + str(50) + '_ms'
#################################
folder_dir = os.listdir('C:/Users/bioel/PycharmProjects/untitled4/' + folder_to_plot)

wb = Workbook()
SVM_acc0 = np.empty(len(folder_dir))
SNN_acc0 = np.empty(len(folder_dir))
DNN_acc0 = np.empty(len(folder_dir))

for single_session in range(len(folder_dir)):
    one = folder_dir[single_session]

    # change this eventually when excel files are just sessions:
    file = folder_to_plot + '/' + str(one)
    wbx = load_workbook(filename=file)

    SVM_sheet = wbx.get_sheet_by_name('SVM')
    SNN_sheet = wbx.get_sheet_by_name('SNN')
    DNN_sheet = wbx.get_sheet_by_name('DNN')

    cell_SVM = SVM_sheet['B2']
    SVM_acc0[single_session] = cell_SVM.value

    cell_SNN = SNN_sheet['G2']
    SNN_acc0[single_session] = cell_SNN.value

    cell_DNN = DNN_sheet['G2']
    DNN_acc0[single_session] = cell_DNN.value

    continue

##### All V1+LGN #####
folder_to_plot = folders_to_grab[1] + '_' + str(50) + '_ms'
#################################
folder_dir = os.listdir('C:/Users/bioel/PycharmProjects/untitled4/' + folder_to_plot)

wb = Workbook()
SVM_acc1 = np.empty(len(folder_dir))
SNN_acc1 = np.empty(len(folder_dir))
DNN_acc1 = np.empty(len(folder_dir))

for single_session in range(len(folder_dir)):
    one = folder_dir[single_session]

    # change this eventually when excel files are just sessions:
    file = folder_to_plot + '/' + str(one)
    wbx = load_workbook(filename=file)

    SVM_sheet = wbx.get_sheet_by_name('SVM')
    SNN_sheet = wbx.get_sheet_by_name('SNN')
    DNN_sheet = wbx.get_sheet_by_name('DNN')

    cell_SVM = SVM_sheet['B2']
    SVM_acc1[single_session] = cell_SVM.value

    cell_SNN = SNN_sheet['G2']
    SNN_acc1[single_session] = cell_SNN.value

    cell_DNN = DNN_sheet['G2']
    DNN_acc1[single_session] = cell_DNN.value

    continue

##### All Cortical #####
folder_to_plot = folders_to_grab[1] + '_' + str(50) + '_ms'
#################################
folder_dir = os.listdir('C:/Users/bioel/PycharmProjects/untitled4/' + folder_to_plot)

wb = Workbook()
SVM_acc2 = np.empty(len(folder_dir))
SNN_acc2 = np.empty(len(folder_dir))
DNN_acc2 = np.empty(len(folder_dir))

for single_session in range(len(folder_dir)):
    one = folder_dir[single_session]

    # change this eventually when excel files are just sessions:
    file = folder_to_plot + '/' + str(one)
    wbx = load_workbook(filename=file)

    SVM_sheet = wbx.get_sheet_by_name('SVM')
    SNN_sheet = wbx.get_sheet_by_name('SNN')
    DNN_sheet = wbx.get_sheet_by_name('DNN')

    cell_SVM = SVM_sheet['B2']
    SVM_acc2[single_session] = cell_SVM.value

    cell_SNN = SNN_sheet['G2']
    SNN_acc2[single_session] = cell_SNN.value

    cell_DNN = DNN_sheet['G2']
    DNN_acc2[single_session] = cell_DNN.value

    continue

##### All V1+LGN #####
folder_to_plot = folders_to_grab[1] + '_' + str(50) + '_ms'
#################################
folder_dir = os.listdir('C:/Users/bioel/PycharmProjects/untitled4/' + folder_to_plot)

wb = Workbook()
SVM_acc3 = np.empty(len(folder_dir))
SNN_acc3 = np.empty(len(folder_dir))
DNN_acc3 = np.empty(len(folder_dir))

for single_session in range(len(folder_dir)):
    one = folder_dir[single_session]

    # change this eventually when excel files are just sessions:
    file = folder_to_plot + '/' + str(one)
    wbx = load_workbook(filename=file)

    SVM_sheet = wbx.get_sheet_by_name('SVM')
    SNN_sheet = wbx.get_sheet_by_name('SNN')
    DNN_sheet = wbx.get_sheet_by_name('DNN')

    cell_SVM = SVM_sheet['B2']
    SVM_acc3[single_session] = cell_SVM.value

    cell_SNN = SNN_sheet['G2']
    SNN_acc3[single_session] = cell_SNN.value

    cell_DNN = DNN_sheet['G2']
    DNN_acc3[single_session] = cell_DNN.value

    continue


# Just pull them down separate
t, p = scip.ttest_ind(SVM_acc0, SVM_acc1)
print('SVM t: ' + str(t))
print('SVM p: '+ str(p))

t2, p2 = scip.ttest_ind(SNN_acc0, SNN_acc1)
print('SNN t: ' + str(t2))
print('SNN p: '+ str(p2))

t3, p3 = scip.ttest_ind(DNN_acc0,DNN_acc1)
print('DNN t: ' + str(t3))
print('DNN p: '+ str(p3))