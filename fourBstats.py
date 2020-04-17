from openpyxl import Workbook, load_workbook
import os
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.figure as figure
import matplotlib.axes as ax
import scipy.stats as scip


############################## FIGURE 4B ##############################
areas = ['VISp','VISrl','VISam', 'VISl', 'VISpm', 'VISal', 'LGd', 'LP', 'CA1', 'CA3']

##### Initialize means across time bins
SVM_acc_mean = np.empty(len(areas))
SNN_acc_mean = np.empty(len(areas))
DNN_acc_mean = np.empty(len(areas))

##### Initialize stds aross time bins
SVM_acc_sem = np.empty(len(areas))
SNN_acc_sem = np.empty(len(areas))
DNN_acc_sem = np.empty(len(areas))


##### INPUT Required Below: #####
file_header = 'C:/Users/bioel/PycharmProjects/untitled4/Local_anat_figure/'
folder_to_plot = areas[0] + '_' + str(50) + '_ms'
#################################
folder_dir = os.listdir('C:/Users/bioel/PycharmProjects/untitled4/Local_anat_figure/' + folder_to_plot)
wb = Workbook()
SVM_acc = np.empty(len(folder_dir))
SNN_acc = np.empty(len(folder_dir))
DNN_acc = np.empty(len(folder_dir))

for single_session in range(len(folder_dir)):
    one = folder_dir[single_session]
    # change this eventually when excel files are just sessions:
    file = file_header + folder_to_plot + '/' + str(one)
    wbx = load_workbook(filename=file)

    SVM_sheet = wbx.get_sheet_by_name('SVM')
    SNN_sheet = wbx.get_sheet_by_name('SNN')
    DNN_sheet = wbx.get_sheet_by_name('DNN')

    cell_SVM = SVM_sheet['B2']
    SVM_acc[single_session] = cell_SVM.value

    cell_SNN = SNN_sheet['G2']
    SNN_acc[single_session] = cell_SNN.value

    cell_DNN = DNN_sheet['G2']
    DNN_acc[single_session] = cell_DNN.value

    continue

##### INPUT Required Below: #####
file_header = 'C:/Users/bioel/PycharmProjects/untitled4/Local_anat_figure/'
folder_to_plot = areas[5] + '_' + str(50) + '_ms'
#################################
folder_dir = os.listdir('C:/Users/bioel/PycharmProjects/untitled4/Local_anat_figure/' + folder_to_plot)
wb = Workbook()
SVM_acc_al = np.empty(len(folder_dir))
SNN_acc_al = np.empty(len(folder_dir))
DNN_acc_al = np.empty(len(folder_dir))

for single_session in range(len(folder_dir)):
    one = folder_dir[single_session]
    # change this eventually when excel files are just sessions:
    file = file_header + folder_to_plot + '/' + str(one)
    wbx = load_workbook(filename=file)

    SVM_sheet = wbx.get_sheet_by_name('SVM')
    SNN_sheet = wbx.get_sheet_by_name('SNN')
    DNN_sheet = wbx.get_sheet_by_name('DNN')

    cell_SVM = SVM_sheet['B2']
    SVM_acc_al[single_session] = cell_SVM.value

    cell_SNN = SNN_sheet['G2']
    SNN_acc_al[single_session] = cell_SNN.value

    cell_DNN = DNN_sheet['G2']
    DNN_acc_al[single_session] = cell_DNN.value

    continue

#####LGN#####

file_header = 'C:/Users/bioel/PycharmProjects/untitled4/Local_anat_figure/'
folder_to_plot = areas[6] + '_' + str(50) + '_ms'
#################################
folder_dir = os.listdir('C:/Users/bioel/PycharmProjects/untitled4/Local_anat_figure/' + folder_to_plot)
wb = Workbook()
SVM_acc_LGN = np.empty(len(folder_dir))
SNN_acc_LGN = np.empty(len(folder_dir))
DNN_acc_LGN = np.empty(len(folder_dir))

for single_session in range(len(folder_dir)):
    one = folder_dir[single_session]
    # change this eventually when excel files are just sessions:
    file = file_header + folder_to_plot + '/' + str(one)
    wbx = load_workbook(filename=file)

    SVM_sheet = wbx.get_sheet_by_name('SVM')
    SNN_sheet = wbx.get_sheet_by_name('SNN')
    DNN_sheet = wbx.get_sheet_by_name('DNN')

    cell_SVM = SVM_sheet['B2']
    SVM_acc_LGN[single_session] = cell_SVM.value

    cell_SNN = SNN_sheet['G2']
    SNN_acc_LGN[single_session] = cell_SNN.value

    cell_DNN = DNN_sheet['G2']
    DNN_acc_LGN[single_session] = cell_DNN.value

    continue


#####LP#####

file_header = 'C:/Users/bioel/PycharmProjects/untitled4/Local_anat_figure/'
folder_to_plot = areas[7] + '_' + str(50) + '_ms'
#################################
folder_dir = os.listdir('C:/Users/bioel/PycharmProjects/untitled4/Local_anat_figure/' + folder_to_plot)
wb = Workbook()
SVM_acc_LP = np.empty(len(folder_dir))
SNN_acc_LP = np.empty(len(folder_dir))
DNN_acc_LP = np.empty(len(folder_dir))

for single_session in range(len(folder_dir)):
    one = folder_dir[single_session]
    # change this eventually when excel files are just sessions:
    file = file_header + folder_to_plot + '/' + str(one)
    wbx = load_workbook(filename=file)

    SVM_sheet = wbx.get_sheet_by_name('SVM')
    SNN_sheet = wbx.get_sheet_by_name('SNN')
    DNN_sheet = wbx.get_sheet_by_name('DNN')

    cell_SVM = SVM_sheet['B2']
    SVM_acc_LP[single_session] = cell_SVM.value

    cell_SNN = SNN_sheet['G2']
    SNN_acc_LP[single_session] = cell_SNN.value

    cell_DNN = DNN_sheet['G2']
    DNN_acc_LP[single_session] = cell_DNN.value

    continue


print('##### VISp vs LGN #####')
t_p_al, p_p_al = scip.ttest_ind(SVM_acc, SVM_acc_al)
print('SVM p vs al t-value: ' +str(t_p_al))
print('SVM p vs al p-value: ' +str(p_p_al))

t_p_al, p_p_al = scip.ttest_ind(SNN_acc, SNN_acc_al)
print('SNN p vs al t-value: ' +str(t_p_al))
print('SNN p vs al p-value: ' +str(p_p_al))

t_p_al, p_p_al = scip.ttest_ind(DNN_acc, DNN_acc_al)
print('DNN p vs al t-value: ' +str(t_p_al))
print('DNN p vs al p-value: ' +str(p_p_al))

#####
print('##### VISp vs LGN #####')
t_p_al, p_p_al = scip.ttest_ind(SVM_acc, SVM_acc_LGN)
print('SVM p vs al t-value: ' +str(t_p_al))
print('SVM p vs al p-value: ' +str(p_p_al))

t_p_al, p_p_al = scip.ttest_ind(SNN_acc, SNN_acc_LGN)
print('SNN p vs al t-value: ' +str(t_p_al))
print('SNN p vs al p-value: ' +str(p_p_al))

t_p_al, p_p_al = scip.ttest_ind(DNN_acc, DNN_acc_LGN)
print('DNN p vs al t-value: ' +str(t_p_al))
print('DNN p vs al p-value: ' +str(p_p_al))

#####
print('##### LGN vs LP #####')
t_p_al, p_p_al = scip.ttest_ind(SVM_acc_LP, SVM_acc_LGN)
print('SVM p vs al t-value: ' +str(t_p_al))
print('SVM p vs al p-value: ' +str(p_p_al))

t_p_al, p_p_al = scip.ttest_ind(SNN_acc_LP, SNN_acc_LGN)
print('SNN p vs al t-value: ' +str(t_p_al))
print('SNN p vs al p-value: ' +str(p_p_al))

t_p_al, p_p_al = scip.ttest_ind(DNN_acc_LP, DNN_acc_LGN)
print('DNN p vs al t-value: ' +str(t_p_al))
print('DNN p vs al p-value: ' +str(p_p_al))


