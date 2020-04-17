## Plotting
from openpyxl import Workbook, load_workbook
import os
import numpy as np
import matplotlib.pyplot as plt
from matplotlib import rcParams
import matplotlib.figure as figure
import scipy.stats as scip


rcParams['font.family'] = 'serif'
rcParams['font.serif'] = ['Times New Roman']
rcParams.update({'font.size': 15})

############################## FIGURE 1 ##############################
x = [5, 25, 50, 75, 100, 125, 150, 175, 200, 225, 250]

##### Initialize means across time bins
SVM_acc_mean = np.empty(len(x))
SNN_acc_mean = np.empty(len(x))
DNN_acc_mean = np.empty(len(x))

##### Initialize stds aross time bins
SVM_acc_sem = np.empty(len(x))
SNN_acc_sem = np.empty(len(x))
DNN_acc_sem = np.empty(len(x))

for bin in range(len(x)):
    ##### INPUT Required Below: #####
    dir_to_unt = 'C:/Users/bioel/PycharmProjects/untitled4/Time_binning_figure/'
    folder_to_plot = 'All V1_' + str(x[bin]) + '_ms'
    #################################
    folder_dir = os.listdir(dir_to_unt + folder_to_plot)

    wb = Workbook()
    SVM_acc = np.empty(len(folder_dir))
    SNN_acc = np.empty(len(folder_dir))
    DNN_acc = np.empty(len(folder_dir))

    for single_session in range(len(folder_dir)):
        one = folder_dir[single_session]

        # change this eventually when excel files are just sessions:
        file = dir_to_unt +folder_to_plot + '/'+ str(one)
        wbx = load_workbook(filename=file)

        SVM_sheet = wbx.get_sheet_by_name('SVM')
        SNN_sheet = wbx.get_sheet_by_name('SNN')
        DNN_sheet = wbx.get_sheet_by_name('DNN')

        cell_SVM = SVM_sheet['B2']
        SVM_acc[single_session] = cell_SVM.value

        cell_SNN = SNN_sheet['G2']
        SNN_acc[single_session] = cell_SNN.value

        cell_DNN = DNN_sheet['G2']
        DNN_acc[single_session] = cell_DNN.value

        continue

    SVM_acc_mean[bin] = np.mean(SVM_acc)
    SNN_acc_mean[bin] = np.mean(SNN_acc)
    DNN_acc_mean[bin] = np.mean(DNN_acc)

    ####################################

    SVM_acc_sem[bin] = scip.sem(SVM_acc)
    SNN_acc_sem[bin] = scip.sem(SNN_acc)
    DNN_acc_sem[bin] = scip.sem(DNN_acc)

    continue

x_vals = np.array(x)

y1 = SVM_acc_mean*100
y2 = SNN_acc_mean*100
y3 = DNN_acc_mean*100

yerr1 = SVM_acc_sem*100
yerr2 = SNN_acc_sem*100
yerr3 = DNN_acc_sem*100

data = [y1, y2, y3]
data_sem = [yerr1,yerr2,yerr3]
columns = x_vals
rows = ['%d year' % x for x in (100, 75, 50)]

plt.axis([0, 255, 0, 100])
#plt.title('Machine Learning-Based Decoding Accuracies')

colors = plt.cm.BuPu(np.linspace(0.2, 0.7, len(rows)))
ecols = plt.cm.BuPu(np.linspace(0.5,1.0, len(rows)))
n_rows = len(data)
cell_text = []
colors = colors[::-1]

for row in range(n_rows):
    #plt.bar(index, data[row], bar_width, yerr=data_sem[row], bottom=y_offset, color=colors[row])
    #y_offset = y_offset + data[row]
    cell_text.append(['%1.1f' % (x) for x in data[row]])

for row in range(len(data_sem)):
    for col in range(len(data_sem[0])):
        sem_inject = str('%1.1f' % data_sem[row][col])
        cell_text[row][col] = cell_text[row][col] + ' ± ' + sem_inject
        continue
    continue
plt.figure(figsize=(12,9))

plt.errorbar(x_vals,y1,yerr=[yerr1/2, yerr1/2], color=colors[2], ecolor=colors[2], capsize=5.0)
plt.errorbar(x_vals,y2,yerr=[yerr2/2, yerr2/2], color=colors[1], ecolor=colors[1], capsize=5.0)
plt.errorbar(x_vals,y3,yerr=[yerr3/2, yerr3/2], color=colors[0],ecolor=colors[0], capsize=5.0)
cell_text.reverse()

# convert time bin durations to str to append ms
columns_ms = []
for i in range(len(columns)):
    column_txt = str(columns[i]) + ' ms'
    columns_ms = np.append(columns_ms, [column_txt])

the_table = plt.table(cellText=cell_text,
                      rowLabels=['DNN','SNN','SVM'],
                      rowColours=colors,
                      colLabels=columns_ms,
                      cellLoc='center',
                      colWidths=[.0909,.0909,.0909,.0909,.0909,.0909,.0909,.0909,.0909,.0909,.0909],
                      rowLoc='center',
                      colLoc='center',
                      loc='bottom',
                      )

the_table.scale(1,1.3)
the_table.auto_set_font_size(False)
the_table.set_fontsize(15)
#the_table.title('Machine Learning-Based Visual Decoding Accuracies per Time Bin Duration')
#plt.axvline([20,40,60])


font = {'fontname': 'Times New Roman',
        'size': 15}
plt.ylabel('Decoding Accuracy (%)', fontdict=font)
plt.ylim(0, 100,)
plt.xticks([])
#plt.tight_layout()
#plt.legend()
#plt.title('Visual Decoding Accuracies Versus Time Bin Duration For All V1 Units', fontdict=font)
plt.show()
############################## END FIGURE 1 ##############################

