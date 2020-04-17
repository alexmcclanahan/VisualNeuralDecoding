from openpyxl import Workbook, load_workbook
import os
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.figure as figure
from matplotlib import rcParams
import scipy.stats as scip

rcParams['font.family'] = 'serif'
rcParams['font.serif'] = ['Times New Roman']

############################## FIGURE 4B ##############################
areas = ['VISp','VISrl','VISam', 'VISl', 'VISpm', 'VISal', 'LGN', 'LP', 'CA1', 'CA3']

##### Initialize means across time bins
SVM_acc_mean = np.empty(len(areas))
SNN_acc_mean = np.empty(len(areas))
DNN_acc_mean = np.empty(len(areas))

##### Initialize stds aross time bins
SVM_acc_sem = np.empty(len(areas))
SNN_acc_sem = np.empty(len(areas))
DNN_acc_sem = np.empty(len(areas))

for bin in range(len(areas)):
    ##### INPUT Required Below: #####
    file_header = 'C:/Users/bioel/PycharmProjects/untitled4/Local_anat_figure/'
    folder_to_plot = areas[bin] + '_' + str(50) + '_ms'
    #################################
    folder_dir = os.listdir('C:/Users/bioel/PycharmProjects/untitled4/Local_anat_figure/' + folder_to_plot)
    wb = Workbook()
    SVM_acc = np.empty(len(folder_dir))
    SNN_acc = np.empty(len(folder_dir))
    DNN_acc = np.empty(len(folder_dir))

    for single_session in range(len(folder_dir)):
        one = folder_dir[single_session]
        # change this eventually when excel files are just sessions:
        file = file_header + folder_to_plot + '/' + str(one)
        wbx = load_workbook(filename=file)

        SVM_sheet = wbx.get_sheet_by_name('SVM')
        SNN_sheet = wbx.get_sheet_by_name('SNN')
        DNN_sheet = wbx.get_sheet_by_name('DNN')

        cell_SVM = SVM_sheet['B2']
        SVM_acc[single_session] = cell_SVM.value

        cell_SNN = SNN_sheet['G2']
        SNN_acc[single_session] = cell_SNN.value

        cell_DNN = DNN_sheet['G2']
        DNN_acc[single_session] = cell_DNN.value

        continue

    SVM_acc_mean[bin] = np.mean(SVM_acc)
    SNN_acc_mean[bin] = np.mean(SNN_acc)
    DNN_acc_mean[bin] = np.mean(DNN_acc)

    ####################################

    SVM_acc_sem[bin] = scip.sem(SVM_acc)
    SNN_acc_sem[bin] = scip.sem(SNN_acc)
    DNN_acc_sem[bin] = scip.sem(DNN_acc)

    continue

a = SVM_acc_mean*100
b = SNN_acc_mean*100
c = DNN_acc_mean*100

yerr1 = SVM_acc_sem*100
yerr2 = SNN_acc_sem*100
yerr3 = DNN_acc_sem*100

### Up is me ###

data = [a,b,c]
print(data)
data_sem = [yerr1,yerr2,yerr3]
print(data_sem)
columns = areas
rows = ['%d year' % x for x in (100, 75, 50)]
#values = np.arange(0, 100, 10)
#value_increment = 5

# Get some pastel shades for the colors
colors = plt.cm.BuPu(np.linspace(0.2, 0.7, len(rows)))
ecols = plt.cm.BuPu(np.linspace(0.5,1.0, len(rows)))
n_rows = len(data)



# Initialize the vertical-offset for the stacked bar chart.
y_offset = np.zeros(len(columns))
# Plot bars and create text labels for the table
cell_text = []
colors = colors[::-1]


#x = np.arange(len(areas))
bar_width = 0.2
width_between = 0.2
r1 = np.arange(len(a))
r2 = [x + width_between for x in r1]
r3 = [x + width_between for x in r2]

plt.figure(figsize=(12,9))

plt.bar(r1, data[0], bar_width,align='edge', yerr=[data_sem[0]/2,data_sem[0]/2], ecolor=ecols[2], capsize=3.0,bottom=y_offset, color=colors[2])
plt.bar(r2, data[1], bar_width,align='edge', yerr=[data_sem[1]/2,data_sem[1]/2],ecolor=ecols[2],capsize=3.0, bottom=y_offset, color=colors[1])
plt.bar(r3, data[2], bar_width, align='edge',yerr=[data_sem[2]/2,data_sem[2]/2], ecolor=ecols[2],capsize=3.0, bottom=y_offset, color=colors[0])

#plt.bar(r1, data[2]-data[1], bar_width, yerr=data_sem[2], ecolor=ecols[0], bottom=data[1], color=colors[0])
#plt.bar(r2, data[1]-data[0], bar_width, yerr=data_sem[1],ecolor=ecols[1], bottom=data[0], color=colors[1])
#plt.bar(r3, data[0], bar_width, yerr=data_sem[0], ecolor=ecols[2],bottom=y_offset, color=colors[2])


for row in range(n_rows):
    #plt.bar(index, data[row], bar_width, yerr=data_sem[row], bottom=y_offset, color=colors[row])
    y_offset = y_offset + data[row]
    cell_text.append(['%1.1f' % (x) for x in data[row]])
    print(cell_text)
# Reverse colors and text labels to display the last value at the top.

# He He He!
for row in range(len(data_sem)):
    for col in range(len(data_sem[0])):
        sem_inject = str('%1.1f' % data_sem[row][col])
        cell_text[row][col] = cell_text[row][col] + ' ± ' + sem_inject
        continue
    continue


cell_text.reverse()
print(cell_text)


# Add a table at the bottom of the axes
the_table = plt.table(cellText=cell_text,
                      rowLabels=['DNN','SNN','SVM'],
                      rowColours=colors,
                      colLabels=columns,
                        cellLoc='center',
                      colWidths=[.1,.1,.1,.1,.1,.1,.1,.1,.1,.1],
                      rowLoc='center',
                      colLoc='center',
                      loc='bottom')

#plt.yticks(values, ['%d' % val for val in values])
# Adjust layout to make room for the table:
plt.subplots_adjust(left=0.2, bottom=0.2)
the_table.scale(1,2.5)
the_table.auto_set_font_size(False)
the_table.set_fontsize(15)
font = {'fontname': 'Times New Roman',
        'size': 15}
plt.ylabel('Decoding Accuracy (%)', fontdict=font)
plt.xlim(-.2,9.8)
plt.ylim(0,100)
plt.xticks([])
#plt.title('Visual Decoding Accuracies of Individual Subregions (fixed 50 ms time bins)')

### SIGS ###

#first sig
y_max = 50
y_min = 5

plt.annotate("", xy=(0.1, y_max), xycoords='data',
           xytext=(5.1, y_max), textcoords='data',
           arrowprops=dict(arrowstyle="-", ec='#aaaaaa',
                           connectionstyle="bar,fraction=0.04"))

plt.text(2.6, y_max + 4.6, 'n.s.',
       horizontalalignment='center',
       verticalalignment='center')

#second sig
y_max = 55
y_min =10

plt.annotate("", xy=(0.3, y_max), xycoords='data',
           xytext=(5.3, y_max), textcoords='data',
           arrowprops=dict(arrowstyle="-", ec='#aaaaaa',
                           connectionstyle="bar,fraction=0.04"))

plt.text(2.8, y_max + 4, '*',
       horizontalalignment='center',
       verticalalignment='center')

#third sig
y_max = 60
y_min = 20

plt.annotate("", xy=(0.5, y_max), xycoords='data',
           xytext=(5.5, y_max), textcoords='data',
           arrowprops=dict(arrowstyle="-", ec='#aaaaaa',
                           connectionstyle="bar,fraction=0.04"))

plt.text(3.0, y_max + 4, '*',
       horizontalalignment='center',
       verticalalignment='center')

#fourth sig
y_max = 45
y_min = 0

plt.annotate("", xy=(6.1, y_max), xycoords='data',
           xytext=(7.1, y_max), textcoords='data',
           arrowprops=dict(arrowstyle="-", ec='#aaaaaa',
                           connectionstyle="bar,fraction=0.15"))

plt.text(6.6, y_max + 3.25, '*',
       horizontalalignment='center',
       verticalalignment='center')

#fifth sig
y_max = 50
y_min = 5

plt.annotate("", xy=(6.3, y_max), xycoords='data',
           xytext=(7.3, y_max), textcoords='data',
           arrowprops=dict(arrowstyle="-", ec='#aaaaaa',
                           connectionstyle="bar,fraction=0.15"))

plt.text(6.8, y_max + 3.25, '**',
       horizontalalignment='center',
       verticalalignment='center')

#sixth sig
y_max = 55
y_min = 10

plt.annotate("", xy=(6.5, y_max), xycoords='data',
           xytext=(7.5, y_max), textcoords='data',
           arrowprops=dict(arrowstyle="-", ec='#aaaaaa',
                           connectionstyle="bar,fraction=0.15"))

plt.text(7, y_max + 3.25, '**',
       horizontalalignment='center',
       verticalalignment='center')


plt.show()