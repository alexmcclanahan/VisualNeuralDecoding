## Plotting
from openpyxl import Workbook, load_workbook
import os
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.figure as figure
import scipy.stats as scip


############################## FIGURE 1 ##############################
x = [5, 25, 50, 75, 100, 125, 150, 175, 200, 225, 250]




##### 5 ms #####
dir_to_unt = 'C:/Users/bioel/PycharmProjects/untitled4/Time_binning_figure/'
folder_to_plot = 'All V1_' + str(x[0]) + '_ms'
#################################
folder_dir = os.listdir(dir_to_unt + folder_to_plot)

wb = Workbook()
SVM_acc5 = np.empty(len(folder_dir))
SNN_acc5 = np.empty(len(folder_dir))
DNN_acc5 = np.empty(len(folder_dir))

for single_session in range(len(folder_dir)):
    one = folder_dir[single_session]

    # change this eventually when excel files are just sessions:
    file = dir_to_unt +folder_to_plot + '/'+ str(one)
    wbx = load_workbook(filename=file)

    SVM_sheet = wbx.get_sheet_by_name('SVM')
    SNN_sheet = wbx.get_sheet_by_name('SNN')
    DNN_sheet = wbx.get_sheet_by_name('DNN')

    cell_SVM = SVM_sheet['B2']
    SVM_acc5[single_session] = cell_SVM.value

    cell_SNN = SNN_sheet['G2']
    SNN_acc5[single_session] = cell_SNN.value

    cell_DNN = DNN_sheet['G2']
    DNN_acc5[single_session] = cell_DNN.value

    continue

##### 25 ms #####
dir_to_unt = 'C:/Users/bioel/PycharmProjects/untitled4/Time_binning_figure/'
folder_to_plot = 'All V1_' + str(x[1]) + '_ms'
#################################
folder_dir = os.listdir(dir_to_unt + folder_to_plot)

wb = Workbook()
SVM_acc25 = np.empty(len(folder_dir))
SNN_acc25 = np.empty(len(folder_dir))
DNN_acc25 = np.empty(len(folder_dir))

for single_session in range(len(folder_dir)):
    one = folder_dir[single_session]

    # change this eventually when excel files are just sessions:
    file = dir_to_unt +folder_to_plot + '/'+ str(one)
    wbx = load_workbook(filename=file)

    SVM_sheet = wbx.get_sheet_by_name('SVM')
    SNN_sheet = wbx.get_sheet_by_name('SNN')
    DNN_sheet = wbx.get_sheet_by_name('DNN')

    cell_SVM = SVM_sheet['B2']
    SVM_acc25[single_session] = cell_SVM.value

    cell_SNN = SNN_sheet['G2']
    SNN_acc25[single_session] = cell_SNN.value

    cell_DNN = DNN_sheet['G2']
    DNN_acc25[single_session] = cell_DNN.value

    continue


t, p = scip.ttest_ind(SVM_acc5, SVM_acc25)
print(t)
print(p)

t, p = scip.ttest_ind(SNN_acc5, SNN_acc25)
print(t)
print(p)

t, p = scip.ttest_ind(DNN_acc5, DNN_acc25)
print(t)
print(p)
